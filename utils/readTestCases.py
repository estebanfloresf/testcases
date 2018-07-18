from openpyxl import load_workbook
import re
import json

class readFile():
    def __init__(self):
        path = 'C:\\Users\\Esteban.Flores\\Documents\\1 Verndale\\2 Projects\\GE-GeneralElectric\\GE TestCases\\0942-(QA) Course Registration Module.xlsx'
        self.wb = load_workbook(path, data_only=True)
        self.cleanWords =  [
            {"from": "Verify", "to": ""},
            {"from": ":", "to": ""},
            {"from": "On click", "to": "cta"},
            {"from": "On hover", "to": "cta"},
            {"from": "Component", "to": ""},
            {"from": "page displays accordingly in mobile", "to": "mobile/tablet"},
            {"from": "rtf (rich text format)", "to": "verify optional content managed rtf (rich text format)"},
        ]
        self.tagWords =  [
            {"has": "text", "tag": "text"},
            {"has": "hover", "tag": "cta"},
            {"has": "click", "tag": "cta"},
            {"has": "rtf", "tag": "text"},
            {"has": "link", "tag": "link"},
            {"has": "image", "tag": "image"},
         
        ]
        self.final =[]

    def __main__(self):

        for a in self.wb.sheetnames:
            validSheet = re.compile('TC|Mobile')
            # validate expression to see if sheetname is an actual testcase
            if(bool(re.search(validSheet, a))):
                self.readCells(a)

    def readCells(self, sheet):
        item = {
            "component":"",
            "testcases":[]
        }
        # Get Component Name of the sheet
        item['component'] = self.cleanCell(self.wb[str(sheet)].cell(row=1,column=2).value)

        # Make a list of all the description columns
        data = [self.wb[str(sheet)].cell(
            row=i, column=2).value for i in range(13, 150)]
        counter = 0
        for cell in data:
            test = {}
            if(cell != None):
                if('Verify' in cell):
                    # Get testcase of sheet
                    test[str(counter)] = cell.lower()
                    counter+=1
                    # Get tag for each testcase
                    for tag in self.tagWords:
                        if(tag['has'] in cell):
                            test["tag"] = tag['tag']
                        if(item['component']=='mobile/tablet'):
                            test["tag"] = 'mobile'
            if(test != {}):
                item["testcases"].append(test)
        
        self.final.append(item)
        
        with open('data.json', 'w') as outfile:
            json.dump(self.final, outfile)

    def cleanCell(self,cell):
        for word in self.cleanWords:
                cell = cell.replace(word['from'],word['to'])
        cell = cell.lower()
        
        return cell.strip()

if(__name__ == "__main__"):
    app=readFile()
    app.__main__()
